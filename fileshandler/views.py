import os
import uuid
import pandas as pd
import pdfplumber
import tabula
from tabula import read_pdf
from urllib.parse import unquote
from fpdf import FPDF
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from docx2pdf import convert
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from utils import docx2pdfUtil
@csrf_exempt
def pdf2docx(request):
    if request.method != 'POST':
        print("不是post方法")
    else:
        upload_file = request.FILES.get('file_pdf2docx')
        # 生成一个唯一个文件名
        file_name = str(uuid.uuid4())+'.pdf'
        # 保存文件到服务器的某个位置
        file_path = os.path.join(settings.MEDIA_ROOT,file_name)
        fs = FileSystemStorage()
        # pdf2docxUtil.pdf2docxUtil(file_path)
        fs.save(file_path,upload_file)

        # 构建url
        file_url = request.build_absolute_uri(settings.MEDIA_URL+file_name)
        print("file path:",file_path)
        print("file url:",file_url)
        print(type(upload_file))
        return HttpResponse(file_url, upload_file, status=200)


@csrf_exempt
def docx2pdf(request):
    if request.method != 'POST':
        print("不是post方法")
    else:
        upload_file = request.FILES.get('file_docx2pdf')
        # 生成一个唯一个文件名
        file_name = str(uuid.uuid4())+'.docx'
        pdf_file_name = file_name.split('.')[0]+'.pdf'
        # 保存文件到服务器的某个位置
        file_path = os.path.join(settings.MEDIA_ROOT,file_name)
        print("file_path:",file_path)
        fs = FileSystemStorage()
        fs.save(file_path,upload_file)
        # 生成对应的pdf文件
        pdf_file_path = convert_pdf_to_docx(file_path)
        print("pdf文件路径:",unquote(pdf_file_path))

        # 生成对应的pdf文件的url
        pdf_url = request.build_absolute_uri(settings.MEDIA_URL+pdf_file_path)
        # 构建url
        file_url = request.build_absolute_uri(settings.MEDIA_URL+pdf_file_name)
        print("file url:",pdf_url)
        print(type(upload_file))
        return HttpResponse(unquote(pdf_url), upload_file, status=200)

def convert_pdf_to_docx(docx_path):
    # 生成PDF文件的路径
    pdf_path = os.path.splitext(docx_path)[0] + '.pdf'
    # 使用docx2pdf库进行转换
    cv = convert(docx_path,pdf_path)
    return pdf_path


def convert_excel_to_pdf(excel_path):
    # 创建PDF文件
    pdf_path = os.path.splitext(excel_path)[0] + '.pdf'
    # 读取Excel文件
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 创建PDF文件
    pdf = canvas.Canvas(pdf_path)

    # 获取Excel中的行和列数
    max_row = sheet.max_row
    max_col = sheet.max_column

    # 定义PDF的页面大小（可以根据需要调整）
    pdf_width, pdf_height = 600, 800
    pdf.setPageSize((pdf_width, pdf_height))

    # 定义字体和字体大小（可以根据需要调整）
    font_size = 10
    pdfmetrics.registerFont(TTFont('KaiTi', 'simkai.ttf'))
    pdf.setFont("KaiTi", font_size)

    # 将Excel中的内容写入PDF
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            # 获取单元格的值
            cell_value = cell.value
            # 写入PDF
            pdf.drawString(cell.column * 100, pdf_height - cell.row * 20, str(cell_value))

    # 保存PDF文件
    pdf.save()
    return pdf_path

def convert_pdf_to_excel(pdf_path):
    # 创建EXCEL文件
    pdf = pdfplumber.open(pdf_path)
    first_page = pdf.pages[0]
    table = first_page.extract_table()
    excel_path = os.path.splitext(pdf_path)[0] + '.xlsx'
    # 将列表转为df
    table_df = pd.DataFrame(table[1:], columns=table[0])

    # 保存excel
    table_df.to_excel(excel_path)
    return excel_path


@csrf_exempt
def excel2pdf(request):
    if request.method != 'POST':
        print("不是post方法")
    else:
        upload_file = request.FILES.get('file_excel2pdf')
        # 生成一个唯一个文件名
        file_name = str(uuid.uuid4()) + '.xlsx'
        pdf_file_name = file_name.split('.')[0] + '.pdf'
        # 保存文件到服务器的某个位置
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        fs = FileSystemStorage()
        fs.save(file_path, upload_file)
        # 生成对应的pdf文件
        pdf_file_path = convert_excel_to_pdf(file_path)

        # 生成对应的pdf文件的url
        pdf_url = request.build_absolute_uri(settings.MEDIA_URL + pdf_file_path)
        # 构建url
        file_url = request.build_absolute_uri(settings.MEDIA_URL + pdf_file_name)
        return HttpResponse(unquote(pdf_url), upload_file, status=200)

@csrf_exempt
def pdf2excel(request):
    if request.method != 'POST':
        print("不是post方法")
    else:
        upload_file = request.FILES.get('file_pdf2excel')
        # 生成一个唯一个文件名
        file_name = str(uuid.uuid4()) + '.pdf'
        pdf_file_name = file_name.split('.')[0] + '.xlsx'
        # 保存文件到服务器的某个位置
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        fs = FileSystemStorage()
        fs.save(file_path, upload_file)
        # 生成对应的pdf文件
        excel_file_path = convert_pdf_to_excel(file_path)

        # 生成对应的pdf文件的url
        excel_url = request.build_absolute_uri(settings.MEDIA_URL + excel_file_path)
        # 构建url
        file_url = request.build_absolute_uri(settings.MEDIA_URL + pdf_file_name)
        return HttpResponse(unquote(excel_url), upload_file, status=200)
